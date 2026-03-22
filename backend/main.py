import glob as _glob_mod
import os
import json
import logging
import secrets
import sqlite3
import threading
import base64
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from contextlib import asynccontextmanager
from pathlib import Path
import pymupdf
import pymupdf4llm
import openpyxl
from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import StreamingResponse, FileResponse, HTMLResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from openai import OpenAI
from dotenv import load_dotenv

logging.basicConfig(level=logging.INFO, format="%(message)s")
logger = logging.getLogger(__name__)

load_dotenv(Path(__file__).resolve().parents[1] / ".env")

# --- Thread-safe DB access via a lock + connection-per-thread fallback ---
_db_path = Path(__file__).with_name("ffu.db")
_db_lock = threading.Lock()

def get_db() -> sqlite3.Connection:
    """Return the module-level connection, protected by _db_lock for writes."""
    return sqlite3.connect(str(_db_path), check_same_thread=False)

# Main connection — used in request handlers (always under _db_lock for writes)
db = get_db()

client = OpenAI(api_key=os.environ["OPENAI_API_KEY"])
data_dir = Path("data")

# Concurrency limits
_MAX_EXTRACT_WORKERS = min(os.cpu_count() or 4, 8)   # CPU-bound OCR
_MAX_API_CONCURRENT = 10                               # parallel OpenAI calls
_api_sem = threading.Semaphore(_MAX_API_CONCURRENT)

# Prevent concurrent /process calls
_process_lock = threading.Lock()


def extract(path: Path) -> str:
    """Extract text from a PDF or Excel file into a markdown string."""
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls", ".xlsm"):
        return _extract_excel(path)
    return _extract_pdf(path)


def _extract_pdf(path: Path) -> str:
    """Extract PDF to markdown with explicit page markers per page."""
    try:
        pages = pymupdf4llm.to_markdown(
            str(path), ignore_images=True, ignore_graphics=True, page_chunks=True
        )
        parts = []
        for i, page in enumerate(pages, 1):
            text = page.get("text", "").strip() if isinstance(page, dict) else str(page).strip()
            if text:
                parts.append(f"[SIDA {i}]\n{text}")
        return "\n\n".join(parts)
    except Exception as e:
        logging.warning(f"pymupdf4llm failed for {path.name} ({e}), falling back to basic text extraction")
        doc = pymupdf.open(str(path))
        parts = []
        for i, page in enumerate(doc, 1):
            text = page.get_text().strip()
            if text:
                parts.append(f"[SIDA {i}]\n{text}")
        doc.close()
        return "\n\n".join(parts)


def _extract_excel(path: Path) -> str:
    """Convert Excel workbook to markdown tables — one section per sheet.
    Each sheet is treated as a 'page' with [SIDA N] markers for tagging."""
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    sections: list[str] = []
    page_num = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Collect non-empty rows
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                rows.append(cells)

        if not rows:
            continue

        page_num += 1

        # Trim trailing empty columns across all rows
        max_col = max((next((i for i in range(len(r) - 1, -1, -1) if r[i]), -1) + 1) for r in rows)
        rows = [r[:max_col] for r in rows]

        # Build markdown table (first row = header)
        header = rows[0]
        sep = ["---"] * len(header)
        lines = [
            f"[SIDA {page_num}]",
            f"## {sheet_name}",
            "| " + " | ".join(header) + " |",
            "| " + " | ".join(sep)   + " |",
        ]
        for row in rows[1:]:
            # Pad short rows to header width
            padded = row + [""] * (len(header) - len(row))
            lines.append("| " + " | ".join(padded) + " |")

        sections.append("\n".join(lines))

    wb.close()
    return "\n\n".join(sections) or "(empty workbook)"


class ChatRequest(BaseModel):
    message: str = ""
    history: list[dict] = []
    document_ids: list[int] = []


@asynccontextmanager
async def lifespan(app):
    db.execute("CREATE TABLE IF NOT EXISTS documents(id INTEGER PRIMARY KEY, filename TEXT, content TEXT)")
    db.execute("""CREATE TABLE IF NOT EXISTS tags(
        id INTEGER PRIMARY KEY,
        document_id INTEGER NOT NULL,
        category TEXT NOT NULL,
        label TEXT NOT NULL,
        page INTEGER NOT NULL
    )""")
    db.execute("CREATE INDEX IF NOT EXISTS idx_tags_doc ON tags(document_id)")
    db.execute("CREATE INDEX IF NOT EXISTS idx_tags_cat ON tags(category)")
    db.commit()
    yield


app = FastAPI(lifespan=lifespan)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ── Simple password protection (set APP_PASSWORD env var to enable) ────────────
_app_password = os.getenv("APP_PASSWORD", "")

@app.middleware("http")
async def password_gate(request: Request, call_next):
    if not _app_password:
        return await call_next(request)
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Basic "):
        try:
            decoded = base64.b64decode(auth[6:]).decode()
            _, _, pwd = decoded.partition(":")
            if secrets.compare_digest(pwd, _app_password):
                return await call_next(request)
        except Exception:
            pass
    return Response(
        status_code=401,
        headers={"WWW-Authenticate": 'Basic realm="FFU Analyzer"'},
        content="Ange lösenord för att komma åt applikationen.",
    )

# ── Strip /api prefix so frontend /api/X hits backend /X ──────────────────────
@app.middleware("http")
async def strip_api_prefix(request: Request, call_next):
    if request.scope["path"].startswith("/api/"):
        request.scope["path"] = request.scope["path"][4:]
    elif request.scope["path"] == "/api":
        request.scope["path"] = "/"
    return await call_next(request)

# ── Serve frontend static files in production ─────────────────────────────────
_frontend_dist = Path(__file__).resolve().parent.parent / "frontend" / "dist"

@app.get("/documents")
def list_documents():
    rows = db.execute("SELECT id, filename FROM documents ORDER BY filename").fetchall()
    return [{"id": r[0], "filename": r[1]} for r in rows]


@app.get("/documents/{doc_id}")
def get_document(doc_id: int):
    row = db.execute("SELECT id, filename, content FROM documents WHERE id = ?", (doc_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Document not found")
    return {"id": row[0], "filename": row[1], "content": row[2]}


@app.get("/documents/{doc_id}/file")
def get_document_file(doc_id: int):
    row = db.execute("SELECT filename FROM documents WHERE id = ?", (doc_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Document not found")
    escaped = _glob_mod.escape(row[0])
    matches = list(data_dir.rglob(escaped))
    if not matches:
        raise HTTPException(status_code=404, detail="File not found on disk")
    return FileResponse(str(matches[0]), filename=row[0])


@app.post("/process")
def process():
    """Extract all documents and stream progress via SSE."""
    if not _process_lock.acquire(blocking=False):
        return StreamingResponse(
            iter([f"data: {json.dumps({'type': 'error', 'message': 'Bearbetning pågår redan'})}\n\n"]),
            media_type="text/event-stream",
        )

    def generate():
        try:
            with _db_lock:
                db.execute("DELETE FROM documents"); db.execute("DELETE FROM tags"); db.commit()
            pdf_paths   = sorted(data_dir.rglob("*.pdf"))
            excel_paths = sorted(p for ext in ("*.xlsx", "*.xls", "*.xlsm") for p in data_dir.rglob(ext))
            paths = pdf_paths + excel_paths
            total = len(paths)
            extracted = 0

            yield f"data: {json.dumps({'type': 'phase', 'phase': 'extract', 'total': total})}\n\n"

            with ProcessPoolExecutor(max_workers=min(total, _MAX_EXTRACT_WORKERS) or 1) as pool:
                futures = {pool.submit(extract, path): path for path in paths}
                for future in as_completed(futures):
                    path = futures[future]
                    try:
                        content = future.result()
                        with _db_lock:
                            db.execute("INSERT INTO documents(filename, content) VALUES(?, ?)", (path.name, content))
                            db.commit()
                        extracted += 1
                        logger.info(f"Extracted {path.name}")
                        yield f"data: {json.dumps({'type': 'extracted', 'filename': path.name, 'count': extracted, 'total': total})}\n\n"
                    except Exception as e:
                        logger.error(f"Skipping {path.name}: {e}")

            # Phase 2: AI tagging — all documents in parallel
            doc_rows = db.execute("SELECT id, filename, content FROM documents ORDER BY id").fetchall()
            total_docs = len(doc_rows)
            yield f"data: {json.dumps({'type': 'phase', 'phase': 'tagging', 'total': total_docs})}\n\n"

            tagged_count = 0
            with ThreadPoolExecutor(max_workers=min(total_docs, _MAX_API_CONCURRENT) or 1) as pool:
                futures = {
                    pool.submit(_generate_tags, content): (doc_id, filename)
                    for doc_id, filename, content in doc_rows
                }
                for future in as_completed(futures):
                    doc_id, filename = futures[future]
                    tagged_count += 1
                    try:
                        tags = future.result()
                        with _db_lock:
                            for tag in tags:
                                db.execute(
                                    "INSERT INTO tags(document_id, category, label, page) VALUES(?,?,?,?)",
                                    (doc_id, tag["category"], tag["label"], tag["page"]),
                                )
                            db.commit()
                        logger.info(f"Tagged {filename} ({len(tags)} tags)")
                        yield f"data: {json.dumps({'type': 'tagged', 'filename': filename, 'count': tagged_count, 'total': total_docs})}\n\n"
                    except Exception as e:
                        logger.error(f"Tagging failed for {filename}: {e}")

            yield f"data: {json.dumps({'type': 'done', 'count': extracted})}\n\n"
        finally:
            _process_lock.release()

    return StreamingResponse(generate(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


def _annotate_pages(content: str) -> str:
    """Ensure [SIDA N] markers are present. New extractions already include them;
    this handles legacy content that used ----- separators and any other content
    (e.g. Excel) that has no page markers at all."""
    if "[SIDA " in content:
        return content
    raw_pages = content.split("-----")
    if len(raw_pages) > 1:
        return "\n\n".join(
            f"[SIDA {i + 1}]\n{p.strip()}"
            for i, p in enumerate(raw_pages) if p.strip()
        )
    # No markers and no separators — wrap entire content as page 1
    if content.strip():
        return f"[SIDA 1]\n{content}"
    return content


@app.get("/tags")
def get_all_tags():
    rows = db.execute(
        "SELECT t.id, t.document_id, t.category, t.label, t.page, d.filename "
        "FROM tags t JOIN documents d ON d.id = t.document_id "
        "ORDER BY t.category, t.label",
    ).fetchall()
    return [{"id": r[0], "document_id": r[1], "category": r[2], "label": r[3], "page": r[4], "filename": r[5]} for r in rows]


@app.get("/timeline")
def get_timeline():
    """Return timeline events — only important dates (deadlines, decisions, milestones)."""
    import re
    rows = db.execute(
        "SELECT t.label, t.page, d.id, d.filename "
        "FROM tags t JOIN documents d ON d.id = t.document_id "
        "WHERE t.category = 'date' ORDER BY t.label"
    ).fetchall()

    # Skip generic document metadata dates
    _skip = re.compile(
        r"^(datum|rev\.?\s*datum|ritningsdatum|senaste\s*ändring|ändringsdatum|"
        r"dokumentdatum|version\b|rev\.?\s*[a-z0-9]|förfrågningsunderlag|"
        r"af\s+daterad|esa\s+daterad|ref\d|wsp\s|tyréns\s|mur\s)",
        re.IGNORECASE,
    )
    # Skip labels that are just a bare date with no context
    _bare_date = re.compile(r"^\d{4}-\d{1,2}-\d{1,2}$")
    # Skip relative durations without a real calendar date
    _relative_only = re.compile(
        r"^(\d+\s*(dagar|veckor|månader|år)|vid\s|efter\s|före\s|senast\s*\d|minst\s*\d|"
        r"ungefär\s|patrullering|röjning|hållbarhet|%|vite\s)",
        re.IGNORECASE,
    )
    # Must contain an actual date to be placeable on a timeline
    _has_date = re.compile(
        r"\d{4}-\d{1,2}(-\d{1,2})?|"
        r"\d{1,2}[./]\d{1,2}[./]\d{4}|"
        r"(januari|februari|mars|april|maj|juni|juli|augusti|september|oktober|november|december)\s+\d{4}",
        re.IGNORECASE,
    )

    events = []
    seen = set()
    for label, page, doc_id, filename in rows:
        lbl = label.strip()
        if _skip.search(lbl):
            continue
        if _bare_date.match(lbl):
            continue
        if _relative_only.search(lbl):
            continue
        if not _has_date.search(lbl):
            continue
        key = (lbl, doc_id)
        if key in seen:
            continue
        seen.add(key)
        events.append({
            "date": lbl,
            "page": page,
            "document_id": doc_id,
            "filename": filename,
        })

    return events


@app.get("/documents/{doc_id}/tags")
def get_document_tags(doc_id: int):
    row = db.execute("SELECT id FROM documents WHERE id = ?", (doc_id,)).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Document not found")
    tags = db.execute(
        "SELECT id, category, label, page FROM tags WHERE document_id = ? ORDER BY category, page",
        (doc_id,),
    ).fetchall()
    return [{"id": t[0], "category": t[1], "label": t[2], "page": t[3]} for t in tags]


_TAG_SYSTEM_PROMPT = (
    "Du är en erfaren byggadministratör och upphandlingsspecialist som analyserar "
    "svenska FFU-dokument (förfrågningsunderlag), bygghandlingar och anbudsunderlag.\n\n"
    "Din uppgift är att noggrant läsa igenom dokumenttexten och identifiera den mest "
    "affärskritiska informationen. Du ska tagga texten i tre kategorier:\n\n"
    "## Kategorier\n\n"
    "### \"date\" — Viktiga datum och tidsfrister\n"
    "Identifiera datum som påverkar projektet konkret:\n"
    "- Anbudstider och sista inlämningsdatum\n"
    "- Byggstart och färdigställandedatum\n"
    "- Garantitider och garantiperiodens längd\n"
    "- Besiktningsdatum och milstolpar\n"
    "- Tidpunkter för samordning eller etablering\n"
    "Skriv alltid ut det faktiska datumet eller tidsperioden i taggen, "
    "t.ex. \"Anbud senast 2025-06-15\" eller \"Garantitid 5 år\".\n\n"
    "### \"entity\" — Nyckelaktörer och parter\n"
    "Identifiera de viktigaste aktörerna som nämns:\n"
    "- Beställare / byggherre (den som upphandlar)\n"
    "- Projektledare och byggledare\n"
    "- Konsulter (konstruktör, geotekniker, etc.)\n"
    "- Myndigheter och tillsynsorgan\n"
    "- Huvudentreprenör om angiven\n"
    "Inkludera organisationsnamn eller personnamn om de nämns, "
    "t.ex. \"Beställare: Mölndals stad\" eller \"Konsult: WSP\".\n\n"
    "### \"risk\" — Risker, krav och villkor\n"
    "Identifiera klausuler och villkor som är ekonomiskt eller juridiskt viktiga:\n"
    "- Viten och förseningsavgifter (belopp om angivet)\n"
    "- Avvikelser från AMA/MER eller andra standarder\n"
    "- Särskilda försäkringskrav eller säkerhetskrav\n"
    "- Ansvarsbegränsningar eller ansvarsfördelning\n"
    "- ÄTA-hantering och prisreglering\n"
    "- Miljökrav, bullerkrav eller arbetstidsbegränsningar\n"
    "Var specifik, t.ex. \"Vite 15 000 kr/dag\" eller \"Avsteg MER Anläggning kap 4\".\n\n"
    "## Regler för sidnummer (KRITISKT)\n"
    "Texten innehåller markeringar som [SIDA 1], [SIDA 2], [SIDA 3] osv.\n"
    "Varje tagg MÅSTE ha rätt sidnummer. Så här avgör du:\n"
    "1. Leta upp den exakta texten/informationen som taggen baseras på\n"
    "2. Titta på vilken [SIDA N]-markering som kommer NÄRMAST FÖRE den texten\n"
    "3. Använd det N:et som sidnummer\n"
    "GISSA ALDRIG sidnummer. Om du inte kan hitta vilken [SIDA N] informationen "
    "hör till, SKIPPA den taggen helt istället för att gissa sida 1.\n\n"
    "## Övriga regler\n"
    "- Tagga ALLT du hittar i varje kategori — missa ingenting viktigt\n"
    "- Varje tagg ska vara koncis, max 8 ord\n"
    "- Svara ENBART med giltig JSON i detta format:\n"
    "  {\"tags\": [{\"category\": \"date|entity|risk\", \"label\": \"...\", \"page\": N}, ...]}\n"
    "- Ingen annan text utanför JSON-objektet"
)

# Maximum characters per chunk sent to the model
_CHUNK_LIMIT = 50_000


def _subsplit(text: str) -> list[str]:
    """Split oversized text into pieces under _CHUNK_LIMIT at paragraph boundaries."""
    paragraphs = text.split("\n\n")
    chunks: list[str] = []
    current = ""
    for para in paragraphs:
        if current and len(current) + len(para) + 2 > _CHUNK_LIMIT:
            chunks.append(current)
            current = para
        else:
            current = current + "\n\n" + para if current else para
    if current:
        chunks.append(current)
    return chunks


def _split_by_pages(content: str) -> list[str]:
    """Split content into chunks at [SIDA N] boundaries, each under _CHUNK_LIMIT.
    Pages that individually exceed the limit are sub-split at paragraph boundaries."""
    import re
    pages = re.split(r'(?=\[SIDA \d+\])', content)
    pages = [p for p in pages if p.strip()]

    chunks: list[str] = []
    current = ""
    for page in pages:
        if len(page) > _CHUNK_LIMIT:
            # Flush accumulator, then sub-split the oversized page
            if current:
                chunks.append(current)
                current = ""
            chunks.extend(_subsplit(page))
        elif current and len(current) + len(page) > _CHUNK_LIMIT:
            chunks.append(current)
            current = page
        else:
            current += page
    if current:
        chunks.append(current)
    return chunks


def _tag_chunk(chunk: str) -> list[dict]:
    """Send one chunk to GPT and return tags (respects global concurrency limit)."""
    try:
        with _api_sem:
            resp = client.chat.completions.create(
                model="gpt-5.4",
                messages=[
                    {"role": "system", "content": _TAG_SYSTEM_PROMPT},
                    {"role": "user", "content": chunk},
                ],
                response_format={"type": "json_object"},
            )
        data = json.loads(resp.choices[0].message.content)
        return data.get("tags", [])
    except Exception as e:
        logger.error(f"  _tag_chunk failed: {e}")
        return []


_RANK_PROMPT = (
    "Du är en expert på svenska byggentreprenader. Du får en lista med taggar "
    "extraherade från ett upphandlingsdokument. Varje tagg har category, label och page.\n\n"
    "Din uppgift: välj de 5–7 VIKTIGASTE taggarna per kategori (date, entity, risk). "
    "Prioritera information som har störst ekonomisk, juridisk eller praktisk påverkan "
    "för en anbudsgivare. Sprid urvalet över hela dokumentet — inte bara tidiga sidor.\n\n"
    "Svara ENBART med JSON: {\"tags\": [{\"category\": \"...\", \"label\": \"...\", \"page\": N}, ...]}\n"
    "Behåll exakt samma label och page som i indata. Ingen annan text."
)


def _generate_tags(content: str) -> list[dict]:
    """Extract tags from the full document, chunking if needed, then rank."""
    annotated = _annotate_pages(content)
    logger.info(f"  Document length: {len(annotated)} chars (limit {_CHUNK_LIMIT})")

    # Collect all tags (single chunk or multiple — chunks tagged in parallel)
    if len(annotated) <= _CHUNK_LIMIT:
        all_tags = _tag_chunk(annotated)
        logger.info(f"  Single chunk → {len(all_tags)} tags")
    else:
        chunks = _split_by_pages(annotated)
        logger.info(f"  Large document — split into {len(chunks)} chunks")
        all_tags = []
        with ThreadPoolExecutor(max_workers=min(len(chunks), _MAX_API_CONCURRENT)) as pool:
            futures = {pool.submit(_tag_chunk, chunk): i for i, chunk in enumerate(chunks)}
            results: list[list[dict]] = [[] for _ in chunks]
            for future in as_completed(futures):
                i = futures[future]
                chunk_tags = future.result()
                results[i] = chunk_tags
                logger.info(f"    Chunk {i+1}: {len(chunks[i])} chars → {len(chunk_tags)} tags")
        for chunk_tags in results:
            all_tags.extend(chunk_tags)

    logger.info(f"  Total raw tags: {len(all_tags)}")

    # Deduplicate: same category + similar label → keep first occurrence
    seen = set()
    unique = []
    for tag in all_tags:
        key = (tag.get("category", ""), tag.get("label", "").lower().strip())
        if key not in seen:
            seen.add(key)
            unique.append(tag)

    logger.info(f"  After dedup: {len(unique)} unique tags")

    # If already small enough, return as-is
    if len(unique) <= 30:
        return unique

    # Second AI call: rank and select the most important tags
    logger.info(f"  Ranking {len(unique)} tags via AI…")
    try:
        with _api_sem:
            resp = client.chat.completions.create(
                model="gpt-5.4",
                messages=[
                    {"role": "system", "content": _RANK_PROMPT},
                    {"role": "user", "content": json.dumps({"tags": unique}, ensure_ascii=False)},
                ],
                response_format={"type": "json_object"},
            )
        ranked = json.loads(resp.choices[0].message.content).get("tags", [])
        logger.info(f"  Ranked result: {len(ranked)} tags")
        return ranked
    except Exception as e:
        logger.error(f"  Ranking failed: {e} — returning first 7 per category")
        result, counts = [], {}
        for tag in unique:
            cat = tag.get("category", "")
            counts[cat] = counts.get(cat, 0) + 1
            if counts[cat] <= 7:
                result.append(tag)
        return result


@app.post("/chat")
def chat(body: ChatRequest):
    docs = db.execute("SELECT id, filename FROM documents ORDER BY id").fetchall()
    doc_names = {doc_id: name for doc_id, name in docs}

    # Pre-load tagged documents into context
    tagged_ids = body.document_ids
    tagged_content_parts = []
    if tagged_ids:
        for did in tagged_ids:
            row = db.execute("SELECT filename, content FROM documents WHERE id = ?", (did,)).fetchone()
            if row:
                fname, content = row
                annotated = _annotate_pages(content)
                tagged_content_parts.append(f"## Bifogat dokument: {fname}\n\n{annotated}")
                logger.info(f"Chat: pre-loaded document {fname} ({len(annotated)} chars)")

    tagged_section = ""
    if tagged_content_parts:
        tagged_section = (
            "\n\n--- BIFOGADE DOKUMENT (läs noggrant och använd som huvudkälla) ---\n\n"
            + "\n\n".join(tagged_content_parts)
            + "\n\n--- SLUT BIFOGADE DOKUMENT ---\n\n"
        )

    system_prompt = (
        "Du är en FFU-dokumentanalytiker för svenska bygghandlingar och anbudsunderlag.\n"
        "Tillgängliga dokument:\n" +
        "\n".join(f"- {name}" for _, name in docs) +
        "\n\n"
        "ABSOLUT KRAV — KÄLLHÄNVISNINGAR:\n"
        "Du MÅSTE ALLTID inkludera minst en källhänvisning i VARJE svar.\n"
        'Formatet är: 【filnamn, s.N, "citat"】\n'
        'Exempel: 【6.3 Avsteg MER Anläggning 23.pdf, s.2, "garantitid om minst 5 år skall lämnas"】\n'
        "Regler:\n"
        "- Skriv ALDRIG dokumentnamn i fetstil (**) — använd ALLTID 【】-formatet istället.\n"
        "- Citatet MÅSTE vara exakt ordagrant kopierat ur dokumentet — kopiera texten tecken för tecken, "
        "ändra INGA ord, lägg inte till eller ta bort ord, och slå inte ihop text från olika ställen.\n"
        "- Citatet ska vara 5–15 ord.\n"
        "- Använd exakt filnamnet (inklusive filändelsen) som listas ovan.\n"
        "- Om du nämner ett dokument utan att citera, lägg ändå till 【filnamn, s.1】.\n"
        "Använd verktyget read_document för att läsa ett dokuments innehåll när du behöver det."
        + (("\n\nAnvändaren har bifogat dokument nedan. Utgå från dessa som huvudkälla "
            "men använd read_document om du behöver mer information från andra dokument.") if tagged_content_parts else "")
    )
    tools = [{
        "type": "function",
        "function": {
            "name": "read_document",
            "description": "Read one FFU document by database id.",
            "parameters": {
                "type": "object",
                "properties": {"document_id": {"type": "integer"}},
                "required": ["document_id"],
            },
        },
    }]

    def generate():
        sources_read: set[int] = set(tagged_ids)
        user_content = body.message
        if tagged_section:
            user_content = tagged_section + user_content
        messages = [
            {"role": "system", "content": system_prompt},
            *body.history,
            {"role": "user", "content": user_content},
        ]
        try:
            for _ in range(10):
                stream = client.chat.completions.create(
                    model="gpt-5.4",
                    messages=messages,
                    tools=tools,
                    stream=True,
                )

                # Accumulate streamed response
                content = ""
                tool_calls_acc: dict[int, dict] = {}
                for chunk in stream:
                    delta = chunk.choices[0].delta
                    if delta.content:
                        content += delta.content
                        yield f"data: {json.dumps({'type': 'delta', 'text': delta.content})}\n\n"
                    if delta.tool_calls:
                        for tc in delta.tool_calls:
                            slot = tool_calls_acc.setdefault(tc.index, {"id": "", "name": "", "arguments": ""})
                            if tc.id:
                                slot["id"] += tc.id
                            if tc.function and tc.function.name:
                                slot["name"] += tc.function.name
                            if tc.function and tc.function.arguments:
                                slot["arguments"] += tc.function.arguments

                if not tool_calls_acc:
                    source_names = [doc_names[did] for did in sorted(sources_read) if did in doc_names]
                    yield f"data: {json.dumps({'type': 'done', 'sources': source_names})}\n\n"
                    return

                # Add assistant turn with tool calls
                tool_calls_list = [
                    {"id": s["id"], "type": "function", "function": {"name": s["name"], "arguments": s["arguments"]}}
                    for s in tool_calls_acc.values()
                ]
                messages.append({"role": "assistant", "content": content or None, "tool_calls": tool_calls_list})

                # Execute each tool call and append results
                for tc in tool_calls_list:
                    args = json.loads(tc["function"]["arguments"])
                    doc_id = args.get("document_id")
                    if doc_id is not None:
                        sources_read.add(doc_id)
                        yield f"data: {json.dumps({'type': 'reading', 'document': doc_names.get(doc_id, str(doc_id))})}\n\n"
                        row = db.execute("SELECT content FROM documents WHERE id = ?", (doc_id,)).fetchone()
                        doc_content = _annotate_pages(row[0]) if row else "Dokument hittades inte."
                        messages.append({
                            "role": "tool",
                            "tool_call_id": tc["id"],
                            "content": doc_content,
                        })

            source_names = [doc_names[did] for did in sorted(sources_read) if did in doc_names]
            yield f"data: {json.dumps({'type': 'done', 'sources': source_names})}\n\n"
        except Exception as e:
            logger.error(f"Chat error: {e}", exc_info=True)
            yield f"data: {json.dumps({'type': 'error', 'message': 'Ett fel uppstod. Försök igen.'})}\n\n"

    return StreamingResponse(generate(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ── Serve frontend SPA in production ──────────────────────────────────────────
if _frontend_dist.is_dir():
    # Catch-all: serve index.html for SPA routes (must be after all API routes)
    @app.get("/{path:path}")
    async def spa_fallback(path: str):
        file = (_frontend_dist / path).resolve()
        if file.is_file() and str(file).startswith(str(_frontend_dist.resolve())):
            return FileResponse(str(file))
        return FileResponse(str(_frontend_dist / "index.html"))


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
